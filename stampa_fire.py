#!/usr/bin/env python3
"""
STAMPA FIRE - Sysadmin Tools Against Malicious Printer Agon Force Inf Registration Engine
v0.0.2
Installazione automatica di stampanti di rete da configurazione Excel.

Requisiti:
    pip install openpyxl pillow

Avvio:
    python stampa_fire.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import subprocess
import ctypes
import sys
import os
import re
import csv
import json
import logging
import datetime
import webbrowser
import openpyxl
from dataclasses import dataclass
from typing import List, Callable, Optional, Tuple

try:
    from PIL import Image, ImageTk
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

APP_NAME     = "STAMPA FIRE"
APP_SUBTITLE = "Sysadmin Tools Against Malicious Printer Agon Force Inf Registration Engine"
APP_VERSION  = "v0.0.2"
APP_DESC     = ("Tool grafico per l'installazione automatica in batch di stampanti "
                "di rete su Windows 10/11, configurabile da file Excel o CSV.")
APP_AUTORE   = "Giovanni Genna"
APP_GITHUB   = "https://github.com/denvermotel/stampa-fire"
APP_SITO     = "https://denvermotel.github.io/stampa-fire/"
APP_LICENZA      = "GNU General Public License v3.0"
APP_LICENZA_URL  = "https://www.gnu.org/licenses/gpl-3.0.html"


def app_base_dir() -> str:
    """Cartella dei file utente (accanto all'exe/script): es. printers.xlsx, logs/."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_dir() -> str:
    """Cartella delle risorse incluse (assets/).

    In una build onefile PyInstaller estrae i dati in una cartella temporanea
    indicata da `sys._MEIPASS`; da sorgente è la cartella dello script.
    """
    if getattr(sys, "frozen", False):
        return getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


# ─────────────────────────────────────────────────────────────────────────────
#  UAC / privilegi
# ─────────────────────────────────────────────────────────────────────────────

def is_admin() -> bool:
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def elevate_and_restart():
    """Rilancia lo script richiedendo elevazione UAC."""
    script = os.path.abspath(sys.argv[0])
    args   = " ".join(f'"{a}"' for a in sys.argv[1:])
    ctypes.windll.shell32.ShellExecuteW(
        None, "runas", sys.executable, f'"{script}" {args}', None, 1
    )
    sys.exit(0)


# ─────────────────────────────────────────────────────────────────────────────
#  Modello dati
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class PrinterConfig:
    nome_stampante: str
    ip: str
    cartella_driver: str
    file_inf: str
    nome_driver: str
    porta: str = ""
    modello: str = ""
    posizione: str = ""
    commento: str = ""
    abilitata: bool = True

    def __post_init__(self):
        # Genera nome porta automaticamente se non fornito
        if not self.porta:
            self.porta = f"IP_{self.ip.replace('.', '_')}"


# ─────────────────────────────────────────────────────────────────────────────
#  Lettura Excel
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLS = {"nome_stampante", "ip", "cartella_driver", "file_inf"}

# Ordine canonico delle colonne, usato come unica fonte di verità per
# import Excel/CSV, generazione template ed export delle stampanti installate.
COLONNE = [
    "nome_stampante", "ip", "cartella_driver", "file_inf", "nome_driver",
    "porta", "modello", "posizione", "commento", "abilitata",
]

CSV_DELIMITER = ";"  # separatore di default (Excel italiano)


def _riga_a_printerconfig(data: dict) -> Optional[PrinterConfig]:
    """Costruisce un PrinterConfig da un dict riga (chiavi = nomi colonna lower).

    Ritorna None se la riga non ha nome_stampante o ip (riga da saltare).
    Risolve nome_driver dal file .inf se mancante. Condiviso da leggi_excel/leggi_csv.
    """
    if not data.get("nome_stampante") or not data.get("ip"):
        return None

    ab_raw = (data.get("abilitata") or "TRUE").upper()
    abilitata = ab_raw not in ("FALSE", "0", "NO", "N")

    pc = PrinterConfig(
        nome_stampante = data["nome_stampante"],
        ip             = data["ip"],
        cartella_driver= data.get("cartella_driver", ""),
        file_inf       = data.get("file_inf", ""),
        nome_driver    = data.get("nome_driver", ""),
        porta          = data.get("porta", ""),
        modello        = data.get("modello", ""),
        posizione      = data.get("posizione", ""),
        commento       = data.get("commento", ""),
        abilitata      = abilitata,
    )
    # Auto-detect nome_driver dal file .inf se mancante nella configurazione
    if not pc.nome_driver and pc.cartella_driver and pc.file_inf:
        inf_path = os.path.join(pc.cartella_driver, pc.file_inf)
        hint = pc.modello or pc.nome_stampante
        pc.nome_driver = leggi_nome_driver_da_inf(inf_path, hint=hint)
    return pc


def _valida_colonne(headers: List[str]) -> None:
    """Solleva ValueError se mancano colonne obbligatorie."""
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(
            f"Colonne obbligatorie mancanti nel file:\n"
            f"{', '.join(sorted(missing))}\n\n"
            f"Colonne trovate: {', '.join(h for h in headers if not h.startswith('_col'))}"
        )


def leggi_config(path: str) -> List[PrinterConfig]:
    """Carica la configurazione stampanti da .csv o .xlsx in base all'estensione."""
    if path.lower().endswith(".csv"):
        return leggi_csv(path)
    return leggi_excel(path)


def leggi_csv(path: str) -> List[PrinterConfig]:
    """Legge la configurazione stampanti da un file CSV.

    Rileva automaticamente il separatore (`;` o `,`) e gestisce il BOM di Excel
    (encoding utf-8-sig). Stessa validazione e logica di leggi_excel.
    """
    # utf-8-sig rimuove l'eventuale BOM scritto da Excel; latin-1 come fallback
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                testo = f.read()
            break
        except UnicodeDecodeError:
            continue
    else:
        testo = ""

    if not testo.strip():
        return []

    # Auto-rilevamento separatore, fallback a CSV_DELIMITER
    try:
        dialect = csv.Sniffer().sniff(testo[:4096], delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = CSV_DELIMITER

    reader = csv.reader(testo.splitlines(), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []

    headers = [
        (h or "").strip().lower() if (h or "").strip() else f"_col{i}"
        for i, h in enumerate(rows[0])
    ]
    _valida_colonne(headers)

    printers: List[PrinterConfig] = []
    for raw in rows[1:]:
        if not any(v and str(v).strip() for v in raw):
            continue
        data = {}
        for i, v in enumerate(raw):
            if i < len(headers):
                data[headers[i]] = str(v).strip() if v is not None else ""
        pc = _riga_a_printerconfig(data)
        if pc is not None:
            printers.append(pc)

    return printers


def leggi_excel(path: str) -> List[PrinterConfig]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [
        str(v).strip().lower() if v is not None else f"_col{i}"
        for i, v in enumerate(header_row)
    ]

    _valida_colonne(headers)

    printers: List[PrinterConfig] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v for v in row if v is not None and str(v).strip()):
            continue

        data = {}
        for i, v in enumerate(row):
            if i < len(headers):
                data[headers[i]] = str(v).strip() if v is not None else ""

        pc = _riga_a_printerconfig(data)
        if pc is not None:
            printers.append(pc)

    return printers


# ─────────────────────────────────────────────────────────────────────────────
#  Export configurazione stampanti installate
# ─────────────────────────────────────────────────────────────────────────────

def _installate_a_righe(stampanti: List[dict]) -> List[List[str]]:
    """Mappa le stampanti installate sulle colonne canoniche (COLONNE).

    cartella_driver e file_inf restano vuoti: il path del driver non è ricavabile
    dal sistema e va compilato manualmente prima del reimport su un altro PC.
    """
    righe: List[List[str]] = []
    for s in stampanti:
        data = {
            "nome_stampante": s.get("Name", ""),
            "ip":             s.get("HostAddress", ""),
            "cartella_driver": "",
            "file_inf":       "",
            "nome_driver":    s.get("DriverName", ""),
            "porta":          s.get("PortName", ""),
            "modello":        "",
            "posizione":      s.get("Location", ""),
            "commento":       s.get("Comment", ""),
            "abilitata":      "TRUE",
        }
        righe.append([data.get(c, "") for c in COLONNE])
    return righe


def esporta_installate(path: str, stampanti: List[dict]) -> None:
    """Esporta le stampanti installate in .csv o .xlsx (formato da estensione)."""
    righe = _installate_a_righe(stampanti)
    if path.lower().endswith(".xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLONNE)
        for r in righe:
            ws.append(r)
        wb.save(path)
    else:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=CSV_DELIMITER)
            writer.writerow(COLONNE)
            writer.writerows(righe)


# ─────────────────────────────────────────────────────────────────────────────
#  Logica di installazione
# ─────────────────────────────────────────────────────────────────────────────

def ping_ip(ip: str) -> bool:
    try:
        result = subprocess.run(
            ["ping", "-n", "2", "-w", "1000", ip],
            capture_output=True, text=True, timeout=8
        )
        return result.returncode == 0
    except Exception:
        return False


def run_ps(cmd: str) -> Tuple[int, str, str]:
    """Esegue un comando PowerShell. Ritorna (returncode, stdout, stderr)."""
    try:
        result = subprocess.run(
            [
                "powershell",
                "-NonInteractive",
                "-NoProfile",
                "-ExecutionPolicy", "Bypass",
                "-Command", cmd,
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
        return result.returncode, result.stdout.strip(), result.stderr.strip()

    except subprocess.TimeoutExpired as e:
        return 1, "", f"TimeoutExpired: {e}"

    except Exception as e:
        return 1, "", f"Errore esecuzione PowerShell: {e}"


def leggi_nome_driver_da_inf(inf_path: str, hint: str = "") -> str:
    """Estrae il nome del driver stampante da un file .inf Windows.

    I file .inf per stampanti hanno questa struttura:
        [Manufacturer]
        %Company% = DriverName,NTx86,NTamd64

        [DriverName.NTamd64]
        "Brother HL-L2370DN series" = INSTALL_SECTION, USBPRINT\\...

    Se `hint` è fornito (es. nome modello), cerca la riga che lo contiene;
    altrimenti restituisce il primo modello trovato nella sezione architettura.
    Ritorna stringa vuota se non rilevabile.
    """
    try:
        content = ""
        for enc in ("utf-8", "latin-1"):
            try:
                with open(inf_path, "r", encoding=enc, errors="strict") as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        if not content:
            return ""

        # Se abbiamo un hint, cerchiamo direttamente una riga modello che lo contenga.
        # Formato riga: "Nome Modello" = SezInstall, HW_ID, ...
        if hint:
            # Cerca corrispondenza parziale case-insensitive tra le righe modello
            for m in re.finditer(r'^\s*"([^"]+)"\s*=\s*\w+[\w.]*\s*,', content, re.MULTILINE):
                if hint.lower() in m.group(1).lower():
                    return m.group(1).strip()

        # Trova il nome-base della sezione modelli dalla riga in [Manufacturer]
        # Supporta sia "Stringa" = NomeSezione  che  %Variabile% = NomeSezione
        mfg_match = re.search(
            r'^\[Manufacturer\].*?(?:"[^"]*"|%[^%]+%)\s*=\s*(\w+)',
            content, re.MULTILINE | re.IGNORECASE | re.DOTALL
        )
        section_name = mfg_match.group(1) if mfg_match else None

        if section_name:
            # Preferisci sezione NTamd64, poi NT, poi base
            for suffix in (".NTamd64", ".NT", ""):
                sec_re = re.compile(
                    rf'^\[{re.escape(section_name)}{re.escape(suffix)}\]',
                    re.MULTILINE | re.IGNORECASE
                )
                m = sec_re.search(content)
                if m:
                    block_start = m.end()
                    next_sec = re.search(r'^\[', content[block_start:], re.MULTILINE)
                    block = content[block_start: block_start + next_sec.start()] if next_sec else content[block_start:]
                    model = re.search(r'^\s*"([^"]+)"\s*=\s*\w+[\w.]*\s*,', block, re.MULTILINE)
                    if model:
                        return model.group(1).strip()

        # Fallback globale: prima stringa quotata in riga con "= Parola,"
        fallback = re.search(r'^\s*"([^"]+)"\s*=\s*\w+[\w.]*\s*,', content, re.MULTILINE)
        if fallback:
            return fallback.group(1).strip()

    except Exception:
        pass
    return ""


def installa_stampante(
    p: PrinterConfig,
    log_fn: Callable[[str, str], None],
    do_ping: bool,
) -> bool:
    """
    Installa una singola stampante di rete.
    Chiama log_fn(messaggio, tag) per aggiornare la GUI.
    Ritorna True in caso di successo.
    """

    def ok(msg: str):  log_fn(f"    \u2714 {msg}", "ok")
    def fail(msg: str): log_fn(f"    \u2718 {msg}", "err")
    def warn(msg: str): log_fn(f"    \u26a0 {msg}", "warn")
    def step(msg: str): log_fn(f"   \u203a {msg}", "step")

    log_fn(f"\n\u25b6 [{p.nome_stampante}]  IP: {p.ip}", "head")
    if p.modello:
        log_fn(f"   Modello: {p.modello}", "")

    # ── 1. Ping ───────────────────────────────────────────────────────────
    if do_ping:
        step("Test connessione IP...")
        if ping_ip(p.ip):
            ok(f"Ping {p.ip}: raggiungibile")
        else:
            fail(f"Ping {p.ip}: non raggiungibile - installazione saltata")
            return False

    # ── 2. Verifica file INF ──────────────────────────────────────────────
    inf_path = os.path.join(p.cartella_driver, p.file_inf)
    if not os.path.isfile(inf_path):
        fail(f"File INF non trovato: {inf_path}")
        return False

    # ── 3. pnputil /add-driver ────────────────────────────────────────────
    step("Installazione driver (pnputil)...")
    try:
        r = subprocess.run(
            ["pnputil", "/add-driver", inf_path, "/install"],
            capture_output=True, text=True, timeout=120
        )
        if r.returncode in (0, 259, 3010):
            ok(f"Driver INF installato: {p.file_inf}")
            if r.returncode == 259:
                ok("Driver già presente nel sistema")
            if r.returncode == 3010:
                warn("Riavvio consigliato per completare l'installazione del driver")
        else:
            detail = r.stderr.strip() or r.stdout.strip()
            fail(f"pnputil errore (rc={r.returncode}): {detail}")
            return False
    except subprocess.TimeoutExpired:
        fail("pnputil timeout - operazione interrotta")
        return False

    # ── 4. Porta TCP/IP ───────────────────────────────────────────────────
    step("Configurazione porta TCP/IP...")
    rc, _, err = run_ps(
        f'if (-not (Get-PrinterPort -Name "{p.porta}" -ErrorAction SilentlyContinue)) {{'
        f'  Add-PrinterPort -Name "{p.porta}" -PrinterHostAddress "{p.ip}"'
        f'}} else {{ Write-Host "Porta gia esistente" }}'
    )
    if rc == 0:
        ok(f"Porta TCP/IP: {p.porta}")
    else:
        fail(f"Errore configurazione porta: {err}")
        return False

    # ── 5. Add-PrinterDriver ──────────────────────────────────────────────
    step("Registrazione driver in Windows...")
    rc, _, err = run_ps(
        f'if (-not (Get-PrinterDriver -Name "{p.nome_driver}" -ErrorAction SilentlyContinue)) {{'
        f'  Add-PrinterDriver -Name "{p.nome_driver}"'
        f'}} else {{ Write-Host "Driver gia registrato" }}'
    )
    if rc == 0:
        ok(f"Driver registrato: {p.nome_driver}")
    else:
        fail(f"Errore Add-PrinterDriver: {err}")
        return False

    # ── 6. Add-Printer ────────────────────────────────────────────────────
    step("Aggiunta stampante al sistema...")
    # Rimuove la stampante se già esiste (evita errori di duplicato)
    run_ps(f'Remove-Printer -Name "{p.nome_stampante}" -ErrorAction SilentlyContinue')

    ps_cmd = (
        f'Add-Printer -Name "{p.nome_stampante}" '
        f'-DriverName "{p.nome_driver}" '
        f'-PortName "{p.porta}"'
    )
    if p.posizione:
        ps_cmd += f' -Location "{p.posizione}"'
    if p.commento:
        ps_cmd += f' -Comment "{p.commento}"'

    rc, _, err = run_ps(ps_cmd)
    if rc == 0:
        ok(f"Stampante installata: {p.nome_stampante}")
        return True
    else:
        fail(f"Errore Add-Printer: {err}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
#  Logica di disinstallazione / elenco stampanti installate
# ─────────────────────────────────────────────────────────────────────────────

def elenca_stampanti_installate() -> List[dict]:
    """Elenca le stampanti installate nel sistema.

    Ritorna una lista di dict con chiavi:
        Name, DriverName, PortName, HostAddress, Location, Comment
    (HostAddress è l'IP della porta TCP/IP, se applicabile).
    Lista vuota in caso di errore o assenza di stampanti.
    """
    cmd = (
        "Get-Printer | ForEach-Object { "
        "$port = Get-PrinterPort -Name $_.PortName -ErrorAction SilentlyContinue; "
        "[PSCustomObject]@{ "
        "Name=$_.Name; DriverName=$_.DriverName; PortName=$_.PortName; "
        "HostAddress=$port.PrinterHostAddress; Location=$_.Location; Comment=$_.Comment "
        "} } | ConvertTo-Json -Depth 3"
    )
    rc, out, _ = run_ps(cmd)
    if rc != 0 or not out:
        return []
    try:
        dati = json.loads(out)
    except (json.JSONDecodeError, ValueError):
        return []

    # ConvertTo-Json restituisce un oggetto singolo (non una lista) se c'è una sola voce
    if isinstance(dati, dict):
        dati = [dati]

    result: List[dict] = []
    for d in dati:
        if not isinstance(d, dict):
            continue
        result.append({
            "Name":        (d.get("Name") or "").strip(),
            "DriverName":  (d.get("DriverName") or "").strip(),
            "PortName":    (d.get("PortName") or "").strip(),
            "HostAddress": (d.get("HostAddress") or "").strip(),
            "Location":    (d.get("Location") or "").strip(),
            "Comment":     (d.get("Comment") or "").strip(),
        })
    return result


def is_stampante_di_rete(info: dict) -> bool:
    """True se la stampante usa una porta TCP/IP di rete.

    Esclude di fatto le stampanti virtuali (PDF, XPS, OneNote, Fax, porte
    PORTPROMPT:/nul: ecc.) che non hanno un host address né una porta IP_.
    """
    if info.get("HostAddress"):
        return True
    return info.get("PortName", "").startswith("IP_")


def disinstalla_stampante(
    info: dict,
    rimuovi_porta: bool,
    rimuovi_driver: bool,
    log_fn: Callable[[str, str], None],
) -> bool:
    """Disinstalla una stampante e, opzionalmente, la sua porta e/o il driver.

    Porta e driver vengono rimossi solo se non più usati da altre stampanti.
    Chiama log_fn(messaggio, tag). Ritorna True se la stampante è stata rimossa.
    """

    def ok(msg: str):   log_fn(f"    ✔ {msg}", "ok")
    def fail(msg: str): log_fn(f"    ✘ {msg}", "err")
    def warn(msg: str): log_fn(f"    ⚠ {msg}", "warn")
    def step(msg: str): log_fn(f"   › {msg}", "step")

    nome   = info.get("Name", "")
    porta  = info.get("PortName", "")
    driver = info.get("DriverName", "")

    log_fn(f"\n▶ [{nome}]", "head")

    # ── 1. Rimozione stampante ────────────────────────────────────────────
    step("Rimozione stampante...")
    rc, _, err = run_ps(f'Remove-Printer -Name "{nome}" -ErrorAction Stop')
    if rc == 0:
        ok(f"Stampante rimossa: {nome}")
    else:
        fail(f"Errore Remove-Printer: {err}")
        return False

    # ── 2. Rimozione porta (solo se non più in uso) ───────────────────────
    if rimuovi_porta and porta:
        step("Rimozione porta TCP/IP...")
        rc, out, err = run_ps(
            f'if (-not (Get-Printer | Where-Object {{ $_.PortName -eq "{porta}" }})) {{'
            f'  Remove-PrinterPort -Name "{porta}" -ErrorAction Stop;'
            f'  Write-Host "RIMOSSA"'
            f'}} else {{ Write-Host "IN_USO" }}'
        )
        if rc != 0:
            warn(f"Porta non rimossa ({porta}): {err}")
        elif "IN_USO" in out:
            warn(f"Porta ancora in uso da altre stampanti, non rimossa: {porta}")
        else:
            ok(f"Porta rimossa: {porta}")

    # ── 3. Rimozione driver (solo se non più in uso) ──────────────────────
    if rimuovi_driver and driver:
        step("Rimozione driver...")
        rc, out, err = run_ps(
            f'if (-not (Get-Printer | Where-Object {{ $_.DriverName -eq "{driver}" }})) {{'
            f'  Remove-PrinterDriver -Name "{driver}" -ErrorAction Stop;'
            f'  Write-Host "RIMOSSO"'
            f'}} else {{ Write-Host "IN_USO" }}'
        )
        if rc != 0:
            warn(f"Driver non rimosso ({driver}): {err}")
        elif "IN_USO" in out:
            warn(f"Driver ancora in uso da altre stampanti, non rimosso: {driver}")
        else:
            ok(f"Driver rimosso: {driver}")

    return True


# ─────────────────────────────────────────────────────────────────────────────
#  Setup logger su file
# ─────────────────────────────────────────────────────────────────────────────

def setup_logger(log_dir: str) -> Tuple[logging.Logger, str]:
    os.makedirs(log_dir, exist_ok=True)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"stampafire_{ts}.log")

    logger = logging.getLogger(f"stampafire_{ts}")
    logger.setLevel(logging.DEBUG)

    fh  = logging.FileHandler(log_file, encoding="utf-8")
    fmt = logging.Formatter("%(asctime)s %(message)s", "%Y-%m-%d %H:%M:%S")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger, log_file


# ─────────────────────────────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────────────────────────────

class StampaFireApp:

    # Palette colori
    C_BG         = "#F0F4F8"
    C_CARD       = "#FFFFFF"
    C_HEADER_BG  = "#1A2B4A"
    C_ACCENT     = "#2563EB"
    C_ACCENT_HOV = "#1D4ED8"
    C_BORDER     = "#CBD5E1"
    C_LABEL      = "#334155"
    C_SUBLABEL   = "#64748B"
    C_STRIPE     = "#F8FAFC"
    C_LOG_BG     = "#1E1E2E"
    C_LOG_FG     = "#CDD6F4"

    def __init__(self, root: tk.Tk, splash: Optional[tk.Toplevel] = None):
        self.root        = root
        self._splash     = splash
        self.printers:   List[PrinterConfig]  = []
        self.check_vars: List[tk.BooleanVar]  = []
        self.running     = False
        self.logger:     Optional[logging.Logger] = None
        self.log_file:   Optional[str]            = None
        self._mgmt_win:  Optional[tk.Toplevel]    = None
        self._info_win:  Optional[tk.Toplevel]    = None

        self.root.title(f"{APP_NAME} - {APP_VERSION}")
        self.root.geometry("840x740")
        self.root.minsize(700, 600)
        self.root.configure(bg=self.C_BG)
        self._setup_window_icon()

        self._setup_styles()
        self._build_ui()
        self._center_window()
        # Carica automaticamente il file Excel predefinito (stessa cartella dell'eseguibile/script).
        # Il caricamento gira in un thread per non bloccare la UI: lo splash resta
        # visibile e reattivo finché il file (anche corposo) non è stato letto.
        default_path = self._find_default_excel()
        if default_path:
            self.file_var.set(default_path)
            self._avvia_caricamento_iniziale(default_path)
        else:
            self._mostra_finestra_principale()

    def _avvia_caricamento_iniziale(self, path: str):
        # Il worker legge il file in background e deposita l'esito in una coda;
        # il main thread la interroga via `after` (polling). Niente chiamate Tk
        # dal thread: evita le race del cross-thread `after` con lo splash.
        q: "queue.Queue" = queue.Queue()

        def worker():
            try:
                q.put((leggi_config(path), None))
            except Exception as e:
                q.put(([], e))

        threading.Thread(target=worker, daemon=True).start()
        self.root.after(60, self._poll_caricamento_iniziale, path, q)

    def _poll_caricamento_iniziale(self, path: str, q: "queue.Queue"):
        try:
            printers, err = q.get_nowait()
        except queue.Empty:
            self.root.after(60, self._poll_caricamento_iniziale, path, q)
            return
        self._completa_caricamento_iniziale(path, printers, err)

    def _completa_caricamento_iniziale(self, path, printers, err):
        if err is None and printers:
            self.printers = printers
            self._aggiorna_lista()
            self._log(
                f"\U0001f4c2 File caricato: {os.path.basename(path)} "
                f"- {len(self.printers)} stampanti trovate",
                tag="info"
            )

        # Mostra la finestra principale prima di eventuali dialog (lo splash è topmost)
        self._mostra_finestra_principale()

        if err is not None:
            messagebox.showerror("Errore lettura file", str(err), parent=self.root)
        elif not printers:
            messagebox.showwarning("File vuoto",
                "Nessuna stampante trovata nel file.\n"
                "Controlla che le righe siano complete.",
                parent=self.root)

    def _mostra_finestra_principale(self):
        """Chiude lo splash (se presente) e mostra la finestra principale."""
        if self._splash is not None:
            try:
                self._splash.destroy()
            except tk.TclError:
                pass
            self._splash = None
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()

    # ── Stili ttk ────────────────────────────────────────────────────────────

    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use("clam")

        s.configure("TFrame",     background=self.C_BG)
        s.configure("Card.TFrame",background=self.C_CARD)
        s.configure("TLabel",     background=self.C_BG, foreground=self.C_LABEL,
                    font=("Segoe UI", 10))
        s.configure("TCheckbutton", background=self.C_CARD, foreground=self.C_LABEL,
                    font=("Segoe UI", 10))
        s.map("TCheckbutton", background=[("active", self.C_CARD)])
        s.configure("TProgressbar", troughcolor=self.C_BORDER,
                    background=self.C_ACCENT, thickness=14)

    # ── Costruzione layout ────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()

        # Contenitore principale con padding
        main = tk.Frame(self.root, bg=self.C_BG)
        main.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)

        self._build_file_card(main)
        self._build_list_card(main)
        self._build_options_card(main)
        self._build_log_card(main)

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=self.C_HEADER_BG, pady=17)
        hdr.pack(fill=tk.X)

        _icon_path = self._get_icon_path()
        if _PIL_OK and _icon_path:
            _img = Image.open(_icon_path).resize((52, 52), Image.LANCZOS)
            self._header_photo = ImageTk.PhotoImage(_img)
            tk.Label(hdr, image=self._header_photo,
                     bg=self.C_HEADER_BG).pack(side=tk.LEFT, padx=(16, 6))
        else:
            tk.Label(hdr, text="\U0001f5a8", font=("Segoe UI Emoji", 22),
                     bg=self.C_HEADER_BG, fg="white").pack(side=tk.LEFT, padx=(16, 6))

        title_f = tk.Frame(hdr, bg=self.C_HEADER_BG)
        title_f.pack(side=tk.LEFT)
        tk.Label(title_f, text=APP_NAME,    font=("Segoe UI", 14, "bold"),
                 bg=self.C_HEADER_BG, fg="white").pack(anchor="w")
        tk.Label(title_f, text=APP_SUBTITLE, font=("Segoe UI", 7),
                 bg=self.C_HEADER_BG, fg="#64748B").pack(anchor="w")
        tk.Label(title_f, text=APP_VERSION, font=("Segoe UI", 9),
                 bg=self.C_HEADER_BG, fg="#94A3B8").pack(anchor="w")

        # Colonna destra: badge amministratore + pulsante Informazioni sotto
        right = tk.Frame(hdr, bg=self.C_HEADER_BG)
        right.pack(side=tk.RIGHT, padx=16)

        if is_admin():
            badge_txt, badge_fg = "\u2714 AMMINISTRATORE", "#4ADE80"
        else:
            badge_txt, badge_fg = "\u26a0 NON AMMINISTRATORE", "#FCA5A5"
        tk.Label(right, text=badge_txt, font=("Segoe UI", 8, "bold"),
                 bg=self.C_HEADER_BG, fg=badge_fg).pack(anchor="e")

        tk.Button(right, text="\u2139  Informazioni", font=("Segoe UI", 9, "bold"),
                  command=self._apri_info,
                  bg=self.C_ACCENT, fg="white",
                  activebackground=self.C_ACCENT_HOV, activeforeground="white",
                  relief="flat", padx=14, pady=6, cursor="hand2"
                  ).pack(anchor="e", pady=(8, 0))

    def _card(self, parent, title: str = "") -> tk.Frame:
        """Crea un pannello 'card' con bordo e titolo opzionale."""
        outer = tk.Frame(parent, bg=self.C_CARD,
                         highlightbackground=self.C_BORDER,
                         highlightthickness=1)
        outer.pack(fill=tk.X, pady=(0, 8))
        if title:
            hdr = tk.Frame(outer, bg="#F1F5F9")
            hdr.pack(fill=tk.X)
            tk.Label(hdr, text=title, font=("Segoe UI", 9, "bold"),
                     bg="#F1F5F9", fg=self.C_SUBLABEL,
                     padx=12, pady=6).pack(side=tk.LEFT)
            self._last_card_header = hdr
        inner = tk.Frame(outer, bg=self.C_CARD)
        inner.pack(fill=tk.X, padx=12, pady=8)
        return inner

    def _build_file_card(self, parent):
        inner = self._card(parent, "File configurazione stampanti (.xlsx / .csv)")

        row = tk.Frame(inner, bg=self.C_CARD)
        row.pack(fill=tk.X)

        self.file_var = tk.StringVar(value="Nessun file selezionato")
        tk.Entry(row, textvariable=self.file_var, state="readonly",
                 font=("Segoe UI", 10), readonlybackground="#F8FAFC",
                 relief="solid", bd=1, fg=self.C_LABEL,
                 disabledforeground=self.C_SUBLABEL
                 ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)

        tk.Button(row, text="\U0001f4c2  Sfoglia",
                  command=self._scegli_file,
                  font=("Segoe UI", 9), bg=self.C_ACCENT, fg="white",
                  activebackground=self.C_ACCENT_HOV, activeforeground="white",
                  relief="flat", padx=12, pady=5, cursor="hand2"
                  ).pack(side=tk.LEFT, padx=(8, 0))

        tk.Button(row, text="\U0001f5a8  Stampanti installate",
                  command=self._apri_gestione_installate,
                  font=("Segoe UI", 9), bg="#475569", fg="white",
                  activebackground="#334155", activeforeground="white",
                  relief="flat", padx=12, pady=5, cursor="hand2"
                  ).pack(side=tk.LEFT, padx=(6, 0))

    def _build_list_card(self, parent):
        # Card con header custom (serve il frame header per i bottoni Tutte/Nessuna)
        outer = tk.Frame(parent, bg=self.C_CARD,
                         highlightbackground=self.C_BORDER,
                         highlightthickness=1)
        outer.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        hdr = tk.Frame(outer, bg="#F1F5F9")
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="Stampanti disponibili",
                 font=("Segoe UI", 9, "bold"),
                 bg="#F1F5F9", fg=self.C_SUBLABEL,
                 padx=12, pady=6).pack(side=tk.LEFT)

        btn_f = tk.Frame(hdr, bg="#F1F5F9")
        btn_f.pack(side=tk.RIGHT, padx=8, pady=4)
        for txt, cmd in [("Tutte", self._select_all), ("Nessuna", self._select_none)]:
            tk.Button(btn_f, text=txt, command=cmd,
                      font=("Segoe UI", 8), bg="#E2E8F0", fg=self.C_LABEL,
                      activebackground="#CBD5E1", relief="flat",
                      padx=10, pady=2, cursor="hand2"
                      ).pack(side=tk.LEFT, padx=2)

        # Canvas scorrevole
        container = tk.Frame(outer, bg=self.C_CARD)
        container.pack(fill=tk.BOTH, expand=True)

        self.list_canvas = tk.Canvas(container, bg=self.C_CARD,
                                     highlightthickness=0, height=180)
        vsb = ttk.Scrollbar(container, orient="vertical",
                            command=self.list_canvas.yview)
        self.scroll_frame = tk.Frame(self.list_canvas, bg=self.C_CARD)

        self.scroll_frame.bind(
            "<Configure>",
            lambda e: self.list_canvas.configure(
                scrollregion=self.list_canvas.bbox("all"))
        )
        self._cw = self.list_canvas.create_window(
            (0, 0), window=self.scroll_frame, anchor="nw"
        )
        self.list_canvas.bind(
            "<Configure>",
            lambda e: self.list_canvas.itemconfig(self._cw, width=e.width)
        )
        self.list_canvas.configure(yscrollcommand=vsb.set)
        self.list_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Scroll con rotella mouse
        self.list_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self.list_canvas.yview_scroll(-1*(e.delta//120), "units")
        )

        # Placeholder iniziale
        self._placeholder = tk.Label(
            self.scroll_frame,
            text="Carica un file Excel per visualizzare le stampanti disponibili",
            font=("Segoe UI", 10), bg=self.C_CARD, fg="#94A3B8"
        )
        self._placeholder.pack(pady=40)

    def _build_options_card(self, parent):
        outer = tk.Frame(parent, bg=self.C_CARD,
                         highlightbackground=self.C_BORDER,
                         highlightthickness=1)
        outer.pack(fill=tk.X, pady=(0, 8))

        opt_inner = tk.Frame(outer, bg=self.C_CARD)
        opt_inner.pack(fill=tk.X, padx=12, pady=8)

        self.ping_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_inner,
            text="Testa connessione IP (ping) prima di ogni installazione",
            variable=self.ping_var
        ).pack(anchor="w")

        # Progress bar
        pb_frame = tk.Frame(outer, bg=self.C_CARD)
        pb_frame.pack(fill=tk.X, padx=12, pady=(0, 4))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(
            pb_frame, variable=self.progress_var,
            maximum=100, mode="determinate"
        )
        self.progress_bar.pack(fill=tk.X)

        self.progress_lbl = tk.Label(
            pb_frame, text="", font=("Segoe UI", 8),
            bg=self.C_CARD, fg=self.C_SUBLABEL
        )
        self.progress_lbl.pack(anchor="e", pady=(2, 0))

        # Bottone installa
        self.install_btn = tk.Button(
            outer,
            text="\u25b6  INSTALLA SELEZIONATE",
            command=self._avvia_installazione,
            font=("Segoe UI", 11, "bold"),
            bg=self.C_ACCENT, fg="white",
            activebackground=self.C_ACCENT_HOV, activeforeground="white",
            relief="flat", padx=20, pady=10, cursor="hand2"
        )
        self.install_btn.pack(pady=(4, 12), padx=12)

    def _build_log_card(self, parent):
        outer = tk.Frame(parent, bg=self.C_CARD,
                         highlightbackground=self.C_BORDER,
                         highlightthickness=1)
        outer.pack(fill=tk.BOTH, expand=True)

        log_hdr = tk.Frame(outer, bg="#1E1E2E")
        log_hdr.pack(fill=tk.X)
        tk.Label(log_hdr, text="\U0001f4cb Log operazioni",
                 font=("Segoe UI", 9, "bold"),
                 bg="#1E1E2E", fg="#CDD6F4",
                 padx=12, pady=6).pack(side=tk.LEFT)

        self.open_log_btn = tk.Button(
            log_hdr, text="Apri file log",
            command=self._apri_log,
            font=("Segoe UI", 8), bg="#313244", fg="#CDD6F4",
            activebackground="#45475A", activeforeground="white",
            relief="flat", padx=8, pady=3, cursor="hand2",
            state="disabled"
        )
        self.open_log_btn.pack(side=tk.RIGHT, padx=8, pady=3)

        self.log_text = scrolledtext.ScrolledText(
            outer, height=9, state="disabled",
            font=("Consolas", 9), bg=self.C_LOG_BG, fg=self.C_LOG_FG,
            insertbackground=self.C_LOG_FG, relief="flat", bd=0, wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # Tag colori log
        self.log_text.tag_configure("ok",   foreground="#A6E3A1")
        self.log_text.tag_configure("err",  foreground="#F38BA8")
        self.log_text.tag_configure("warn", foreground="#F9E2AF")
        self.log_text.tag_configure("step", foreground="#89B4FA")
        self.log_text.tag_configure("head", foreground="#CBA6F7",
                                    font=("Consolas", 9, "bold"))
        self.log_text.tag_configure("info", foreground="#74C7EC")

    # ── Utilities ─────────────────────────────────────────────────────────────

    def _center_window(self):
        self.root.update_idletasks()
        w  = self.root.winfo_width()
        h  = self.root.winfo_height()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _get_app_dir(self) -> str:
        return app_base_dir()

    def _get_icon_path(self) -> Optional[str]:
        # Risorsa inclusa nel bundle: usa resource_dir (gestisce _MEIPASS in onefile)
        path = os.path.join(resource_dir(), "assets", "icona.png")
        return path if os.path.isfile(path) else None

    def _setup_window_icon(self):
        if not _PIL_OK:
            return
        icon_path = self._get_icon_path()
        if not icon_path:
            return
        img = Image.open(icon_path).resize((32, 32), Image.LANCZOS)
        self._icon_photo = ImageTk.PhotoImage(img)
        self.root.iconphoto(True, self._icon_photo)

    def _find_default_excel(self) -> Optional[str]:
        app_dir = self._get_app_dir()
        candidates = [
            os.path.join(app_dir, "printers.xlsx"),
            os.path.join(app_dir, "printers_template.xlsx"),
            os.path.join(app_dir, "stampanti.xlsx"),
        ]

        for candidate in candidates:
            if os.path.isfile(candidate):
                return candidate

        # fallback al primo .xlsx nella cartella, se presente
        for fn in os.listdir(app_dir):
            if fn.lower().endswith(".xlsx"):
                return os.path.join(app_dir, fn)

        return None

    def _log(self, msg: str, tag: str = ""):
        """Thread-safe: scrive un messaggio nel log widget e nel file."""
        def _do():
            self.log_text.config(state="normal")
            # Auto-tag se non specificato
            t = tag
            if not t:
                if "\u2714" in msg:  t = "ok"
                elif "\u2718" in msg: t = "err"
                elif "\u26a0" in msg: t = "warn"
            self.log_text.insert(tk.END, msg + "\n", t)
            self.log_text.see(tk.END)
            self.log_text.config(state="disabled")
        self.root.after(0, _do)

        if self.logger:
            # Rimuove emoji per il file di log (ASCII-safe)
            clean = msg.encode("ascii", "ignore").decode("ascii").strip()
            self.logger.info(clean)

    # ── Selezione file ────────────────────────────────────────────────────────

    def _scegli_file(self):
        path = filedialog.askopenfilename(
            title="Seleziona file di configurazione stampanti",
            filetypes=[
                ("Excel / CSV", "*.xlsx *.csv"),
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv"),
                ("Tutti i file", "*.*"),
            ]
        )
        if not path:
            return
        self.file_var.set(path)
        self._carica_stampanti(path)

    def _carica_stampanti(self, path: str):
        try:
            self.printers = leggi_config(path)
        except Exception as e:
            messagebox.showerror("Errore lettura file", str(e))
            return

        if not self.printers:
            messagebox.showwarning("File vuoto",
                "Nessuna stampante trovata nel file Excel.\n"
                "Controlla che le righe siano complete.")
            return

        self._aggiorna_lista()
        self._log(
            f"\U0001f4c2 File caricato: {os.path.basename(path)} "
            f"- {len(self.printers)} stampanti trovate",
            tag="info"
        )

    def _aggiorna_lista(self):
        for w in self.scroll_frame.winfo_children():
            w.destroy()
        self.check_vars.clear()

        if not self.printers:
            tk.Label(self.scroll_frame,
                     text="Nessuna stampante trovata",
                     font=("Segoe UI", 10), bg=self.C_CARD, fg="#94A3B8"
                     ).pack(pady=30)
            return

        for i, p in enumerate(self.printers):
            var = tk.BooleanVar(value=p.abilitata)
            self.check_vars.append(var)

            row_bg = self.C_CARD if i % 2 == 0 else self.C_STRIPE
            row = tk.Frame(self.scroll_frame, bg=row_bg)
            row.pack(fill=tk.X, padx=8, pady=6)

            tk.Checkbutton(row, variable=var, bg=row_bg,
                           activebackground=row_bg, cursor="hand2"
                           ).pack(side=tk.LEFT)

            info = tk.Frame(row, bg=row_bg)
            info.pack(side=tk.LEFT, fill=tk.X, expand=True)

            name_txt = p.nome_stampante + (f"  -  {p.modello}" if p.modello else "")
            tk.Label(info, text=name_txt, font=("Segoe UI", 10, "bold"),
                     bg=row_bg, fg="#1E293B", anchor="w").pack(anchor="w")

            detail = f"IP: {p.ip}"
            if p.posizione: detail += f"   \U0001f4cd {p.posizione}"
            if p.commento:  detail += f"   \U0001f4ac {p.commento}"
            tk.Label(info, text=detail, font=("Segoe UI", 8),
                     bg=row_bg, fg=self.C_SUBLABEL, anchor="w").pack(anchor="w")

            # Badge driver
            tk.Label(row, text=p.nome_driver,
                     font=("Segoe UI", 8), bg="#EFF6FF", fg="#2563EB",
                     padx=6, pady=2, relief="flat"
                     ).pack(side=tk.RIGHT, padx=(0, 4))

    def _select_all(self):
        for v in self.check_vars: v.set(True)

    def _select_none(self):
        for v in self.check_vars: v.set(False)

    # ── Installazione ─────────────────────────────────────────────────────────

    def _avvia_installazione(self):
        selected = [(i, p) for i, p in enumerate(self.printers)
                    if i < len(self.check_vars) and self.check_vars[i].get()]

        if not selected:
            messagebox.showwarning("Nessuna selezione",
                "Seleziona almeno una stampante da installare.")
            return

        if not messagebox.askyesno(
            "Conferma installazione",
            f"Verranno installate {len(selected)} stampante/i.\n\nContinuare?"
        ):
            return

        self.running = True
        self.install_btn.config(state="disabled",
                                text="\u23f3  Installazione in corso...")
        self.progress_var.set(0)
        self.progress_lbl.config(text="")

        # Inizializza logger su file
        log_dir = os.path.join(
            os.path.dirname(os.path.abspath(sys.argv[0])), "logs"
        )
        self.logger, self.log_file = setup_logger(log_dir)
        self.open_log_btn.config(state="normal")

        do_ping = self.ping_var.get()

        threading.Thread(
            target=self._run_installation,
            args=(selected, do_ping),
            daemon=True
        ).start()

    def _run_installation(self, selected: list, do_ping: bool):
        total   = len(selected)
        results = []

        ts = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        sep = "\u2550" * 44
        self._log(sep, "head")
        self._log(f"  {APP_NAME} {APP_VERSION}", "head")
        self._log(f"  Avvio: {ts} - {total} stampante/i", "head")
        self._log(sep + "\n", "head")

        for idx, (_, p) in enumerate(selected):
            self.root.after(
                0, self._update_progress, idx, total, p.nome_stampante
            )
            try:
                success = installa_stampante(p, self._log, do_ping)
            except Exception as e:
                success = False
                self._log(
                    f"    \u2718 Errore inaspettato per {p.nome_stampante}: {e}",
                    "err"
                )

            results.append((p.nome_stampante, success))

            esito = "\u2714 OK" if success else "\u2718 FALLITA"
            self._log(f"  {esito}: {p.nome_stampante}",
                      "ok" if success else "err")

        self.root.after(0, self._update_progress, total, total, "")
        self.root.after(0, self._mostra_riepilogo, results)

    def _update_progress(self, done: int, total: int, current: str):
        pct = (done / total * 100) if total else 0
        self.progress_var.set(pct)
        if current:
            self.progress_lbl.config(text=f"{done}/{total}  -  {current}")
        else:
            self.progress_lbl.config(text=f"Completato: {total}/{total}")

    def _mostra_riepilogo(self, results: list):
        self.running = False
        self.install_btn.config(state="normal",
                                text="\u25b6  INSTALLA SELEZIONATE")

        ok_list  = [n for n, s in results if s]
        err_list = [n for n, s in results if not s]

        sep = "\u2550" * 44
        self._log(f"\n{sep}", "head")
        self._log("  RIEPILOGO INSTALLAZIONE", "head")
        self._log(sep, "head")
        self._log(f"  \u2714 Riuscite : {len(ok_list)}/{len(results)}", "ok")
        for n in ok_list:
            self._log(f"     \u2022 {n}", "ok")
        if err_list:
            self._log(f"  \u2718 Fallite  : {len(err_list)}/{len(results)}", "err")
            for n in err_list:
                self._log(f"     \u2022 {n}", "err")
        self._log(f"  Log: {self.log_file}", "info")
        self._log(sep + "\n", "head")

        # Dialogo di riepilogo
        if not err_list:
            messagebox.showinfo(
                "Installazione completata",
                f"\u2714 Tutte le {len(ok_list)} stampanti installate correttamente!"
            )
        elif not ok_list:
            messagebox.showerror(
                "Installazione fallita",
                f"\u2718 Nessuna stampante installata.\n\n"
                f"Verifica il log per i dettagli:\n{self.log_file}"
            )
        else:
            messagebox.showwarning(
                "Installazione parziale",
                f"\u2714 Riuscite : {len(ok_list)}\n"
                f"\u2718 Fallite  : {len(err_list)}\n\n"
                f"Stampanti non installate:\n"
                + "\n".join(f"  \u2022 {n}" for n in err_list)
                + f"\n\nVerifica il log per i dettagli."
            )

    def _apri_log(self):
        if self.log_file and os.path.isfile(self.log_file):
            os.startfile(self.log_file)

    # ── Finestra informazioni ─────────────────────────────────────────────────

    def _apri_info(self):
        # Evita di aprire due finestre informazioni contemporaneamente
        if getattr(self, "_info_win", None) is not None and self._info_win.winfo_exists():
            self._info_win.lift()
            self._info_win.focus_force()
            return

        win = tk.Toplevel(self.root)
        self._info_win = win
        win.title(f"{APP_NAME} - Informazioni")
        win.configure(bg=self.C_CARD)
        win.resizable(False, False)
        if getattr(self, "_icon_photo", None):
            win.iconphoto(False, self._icon_photo)
        win.transient(self.root)
        win.protocol("WM_DELETE_WINDOW",
                     lambda: (win.destroy(), setattr(self, "_info_win", None)))

        # Intestazione scura con icona/nome
        head = tk.Frame(win, bg=self.C_HEADER_BG, pady=16)
        head.pack(fill=tk.X)
        if _PIL_OK and getattr(self, "_header_photo", None):
            tk.Label(head, image=self._header_photo,
                     bg=self.C_HEADER_BG).pack(side=tk.LEFT, padx=(18, 8))
        else:
            tk.Label(head, text="\U0001f5a8", font=("Segoe UI Emoji", 22),
                     bg=self.C_HEADER_BG, fg="white").pack(side=tk.LEFT, padx=(18, 8))
        htxt = tk.Frame(head, bg=self.C_HEADER_BG)
        htxt.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(htxt, text=APP_NAME, font=("Segoe UI", 14, "bold"),
                 bg=self.C_HEADER_BG, fg="white").pack(anchor="w")
        tk.Label(htxt, text=APP_VERSION, font=("Segoe UI", 9),
                 bg=self.C_HEADER_BG, fg="#94A3B8").pack(anchor="w")

        body = tk.Frame(win, bg=self.C_CARD, padx=22, pady=18)
        body.pack(fill=tk.BOTH, expand=True)

        tk.Label(body, text=APP_SUBTITLE, font=("Segoe UI", 8, "italic"),
                 bg=self.C_CARD, fg=self.C_SUBLABEL, wraplength=380,
                 justify="left").pack(anchor="w", pady=(0, 10))
        tk.Label(body, text=APP_DESC, font=("Segoe UI", 10),
                 bg=self.C_CARD, fg=self.C_LABEL, wraplength=380,
                 justify="left").pack(anchor="w", pady=(0, 14))

        def riga(label: str, valore: str):
            r = tk.Frame(body, bg=self.C_CARD)
            r.pack(fill=tk.X, pady=2)
            tk.Label(r, text=label, font=("Segoe UI", 9, "bold"), width=16,
                     anchor="w", bg=self.C_CARD, fg=self.C_SUBLABEL).pack(side=tk.LEFT)
            tk.Label(r, text=valore, font=("Segoe UI", 10),
                     bg=self.C_CARD, fg=self.C_LABEL, anchor="w").pack(side=tk.LEFT)

        def riga_link(label: str, testo: str, url: str):
            r = tk.Frame(body, bg=self.C_CARD)
            r.pack(fill=tk.X, pady=2)
            tk.Label(r, text=label, font=("Segoe UI", 9, "bold"), width=16,
                     anchor="w", bg=self.C_CARD, fg=self.C_SUBLABEL).pack(side=tk.LEFT)
            link = tk.Label(r, text=testo, font=("Segoe UI", 10, "underline"),
                            bg=self.C_CARD, fg=self.C_ACCENT, cursor="hand2", anchor="w")
            link.pack(side=tk.LEFT)
            link.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))

        riga("Autore originale:", APP_AUTORE)
        riga_link("GitHub:", APP_GITHUB, APP_GITHUB)
        riga_link("Guida / sito:", APP_SITO, APP_SITO)
        riga_link("Licenza:", APP_LICENZA, APP_LICENZA_URL)

        tk.Button(body, text="Chiudi",
                  command=lambda: (win.destroy(), setattr(self, "_info_win", None)),
                  font=("Segoe UI", 9), bg=self.C_ACCENT, fg="white",
                  activebackground=self.C_ACCENT_HOV, activeforeground="white",
                  relief="flat", padx=18, pady=6, cursor="hand2"
                  ).pack(anchor="e", pady=(18, 0))

        win.update_idletasks()
        w, h = win.winfo_width(), win.winfo_height()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - h) // 2
        win.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    # ── Gestione stampanti installate ─────────────────────────────────────────

    def _apri_gestione_installate(self):
        # Evita di aprire due finestre contemporaneamente
        if self._mgmt_win is not None and self._mgmt_win.winfo_exists():
            self._mgmt_win.lift()
            self._mgmt_win.focus_force()
            return

        win = tk.Toplevel(self.root)
        self._mgmt_win = win
        win.title(f"{APP_NAME} - Stampanti installate")
        win.geometry("760x620")
        win.minsize(640, 520)
        win.configure(bg=self.C_BG)
        if getattr(self, "_icon_photo", None):
            win.iconphoto(False, self._icon_photo)
        win.transient(self.root)
        win.protocol("WM_DELETE_WINDOW", self._chiudi_gestione_installate)

        self._mgmt_printers: List[dict]          = []
        self._mgmt_check_vars: List[tk.BooleanVar] = []
        self._mgmt_busy = False

        # ── Barra superiore ────────────────────────────────────────────────
        topbar = tk.Frame(win, bg=self.C_CARD,
                          highlightbackground=self.C_BORDER, highlightthickness=1)
        topbar.pack(fill=tk.X, padx=14, pady=(12, 8))
        topbar_in = tk.Frame(topbar, bg=self.C_CARD)
        topbar_in.pack(fill=tk.X, padx=12, pady=8)

        tk.Button(topbar_in, text="\U0001f504  Aggiorna",
                  command=self._aggiorna_lista_installate,
                  font=("Segoe UI", 9), bg=self.C_ACCENT, fg="white",
                  activebackground=self.C_ACCENT_HOV, activeforeground="white",
                  relief="flat", padx=12, pady=5, cursor="hand2"
                  ).pack(side=tk.LEFT)

        tk.Button(topbar_in, text="\U0001f4be  Esporta",
                  command=self._esporta_installate,
                  font=("Segoe UI", 9), bg="#0F766E", fg="white",
                  activebackground="#115E59", activeforeground="white",
                  relief="flat", padx=12, pady=5, cursor="hand2"
                  ).pack(side=tk.LEFT, padx=(8, 0))

        self._mgmt_filtro_rete = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            topbar_in, text="Mostra solo stampanti di rete",
            variable=self._mgmt_filtro_rete,
            command=self._aggiorna_lista_installate
        ).pack(side=tk.LEFT, padx=(12, 0))

        for txt, cmd in [("Tutte", self._mgmt_select_all),
                         ("Nessuna", self._mgmt_select_none)]:
            tk.Button(topbar_in, text=txt, command=cmd,
                      font=("Segoe UI", 8), bg="#E2E8F0", fg=self.C_LABEL,
                      activebackground="#CBD5E1", relief="flat",
                      padx=10, pady=3, cursor="hand2"
                      ).pack(side=tk.RIGHT, padx=2)

        # ── Lista scorrevole ───────────────────────────────────────────────
        list_outer = tk.Frame(win, bg=self.C_CARD,
                              highlightbackground=self.C_BORDER, highlightthickness=1)
        list_outer.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))

        self._mgmt_canvas = tk.Canvas(list_outer, bg=self.C_CARD,
                                      highlightthickness=0)
        vsb = ttk.Scrollbar(list_outer, orient="vertical",
                            command=self._mgmt_canvas.yview)
        self._mgmt_scroll = tk.Frame(self._mgmt_canvas, bg=self.C_CARD)
        self._mgmt_scroll.bind(
            "<Configure>",
            lambda e: self._mgmt_canvas.configure(
                scrollregion=self._mgmt_canvas.bbox("all"))
        )
        cw = self._mgmt_canvas.create_window(
            (0, 0), window=self._mgmt_scroll, anchor="nw")
        self._mgmt_canvas.bind(
            "<Configure>",
            lambda e: self._mgmt_canvas.itemconfig(cw, width=e.width)
        )
        self._mgmt_canvas.configure(yscrollcommand=vsb.set)
        self._mgmt_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # ── Opzioni di rimozione ───────────────────────────────────────────
        opt_outer = tk.Frame(win, bg=self.C_CARD,
                            highlightbackground=self.C_BORDER, highlightthickness=1)
        opt_outer.pack(fill=tk.X, padx=14, pady=(0, 8))
        opt_in = tk.Frame(opt_outer, bg=self.C_CARD)
        opt_in.pack(fill=tk.X, padx=12, pady=8)

        tk.Label(opt_in,
                 text="Di default viene rimossa solo la stampante. Opzioni aggiuntive:",
                 font=("Segoe UI", 8), bg=self.C_CARD, fg=self.C_SUBLABEL
                 ).pack(anchor="w")

        self._mgmt_rm_porta  = tk.BooleanVar(value=False)
        self._mgmt_rm_driver = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_in, text="Rimuovi anche la porta TCP/IP (se non più in uso)",
                        variable=self._mgmt_rm_porta).pack(anchor="w", pady=(4, 0))
        ttk.Checkbutton(opt_in, text="Rimuovi anche il driver (se non più in uso)",
                        variable=self._mgmt_rm_driver).pack(anchor="w")

        self._mgmt_uninstall_btn = tk.Button(
            opt_outer, text="\U0001f5d1  DISINSTALLA SELEZIONATE",
            command=self._avvia_disinstallazione,
            font=("Segoe UI", 11, "bold"),
            bg="#DC2626", fg="white",
            activebackground="#B91C1C", activeforeground="white",
            relief="flat", padx=20, pady=9, cursor="hand2"
        )
        self._mgmt_uninstall_btn.pack(pady=(2, 12), padx=12)

        # ── Log locale ─────────────────────────────────────────────────────
        log_outer = tk.Frame(win, bg=self.C_CARD,
                            highlightbackground=self.C_BORDER, highlightthickness=1)
        log_outer.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 12))

        self._mgmt_log = scrolledtext.ScrolledText(
            log_outer, height=7, state="disabled",
            font=("Consolas", 9), bg=self.C_LOG_BG, fg=self.C_LOG_FG,
            insertbackground=self.C_LOG_FG, relief="flat", bd=0, wrap=tk.WORD
        )
        self._mgmt_log.pack(fill=tk.BOTH, expand=True)
        for tag, color in [("ok", "#A6E3A1"), ("err", "#F38BA8"),
                           ("warn", "#F9E2AF"), ("step", "#89B4FA"),
                           ("info", "#74C7EC")]:
            self._mgmt_log.tag_configure(tag, foreground=color)
        self._mgmt_log.tag_configure("head", foreground="#CBA6F7",
                                     font=("Consolas", 9, "bold"))

        # Caricamento iniziale automatico
        self._aggiorna_lista_installate()

    def _chiudi_gestione_installate(self):
        if self._mgmt_busy:
            if not messagebox.askyesno(
                "Operazione in corso",
                "Una disinstallazione è in corso. Chiudere comunque la finestra?",
                parent=self._mgmt_win
            ):
                return
        if self._mgmt_win is not None:
            self._mgmt_win.destroy()
        self._mgmt_win = None

    def _mgmt_log_write(self, msg: str, tag: str = ""):
        """Thread-safe: scrive nel log della finestra di gestione e nel file."""
        def _do():
            if self._mgmt_win is None or not self._mgmt_win.winfo_exists():
                return
            self._mgmt_log.config(state="normal")
            self._mgmt_log.insert(tk.END, msg + "\n", tag)
            self._mgmt_log.see(tk.END)
            self._mgmt_log.config(state="disabled")
        self.root.after(0, _do)

        if self.logger:
            clean = msg.encode("ascii", "ignore").decode("ascii").strip()
            self.logger.info(clean)

    def _aggiorna_lista_installate(self):
        if self._mgmt_busy:
            return
        self._mostra_placeholder_installate("Caricamento stampanti installate...")

        def worker():
            stampanti = elenca_stampanti_installate()
            self.root.after(0, self._popola_lista_installate, stampanti)

        threading.Thread(target=worker, daemon=True).start()

    def _mostra_placeholder_installate(self, testo: str):
        if self._mgmt_win is None or not self._mgmt_win.winfo_exists():
            return
        for w in self._mgmt_scroll.winfo_children():
            w.destroy()
        self._mgmt_check_vars = []
        tk.Label(self._mgmt_scroll, text=testo,
                 font=("Segoe UI", 10), bg=self.C_CARD, fg="#94A3B8"
                 ).pack(pady=40)

    def _popola_lista_installate(self, stampanti: List[dict]):
        if self._mgmt_win is None or not self._mgmt_win.winfo_exists():
            return

        if self._mgmt_filtro_rete.get():
            stampanti = [s for s in stampanti if is_stampante_di_rete(s)]
        self._mgmt_printers = stampanti

        for w in self._mgmt_scroll.winfo_children():
            w.destroy()
        self._mgmt_check_vars = []

        if not stampanti:
            self._mostra_placeholder_installate(
                "Nessuna stampante trovata. Premi 'Aggiorna' per ricaricare.")
            return

        for i, s in enumerate(stampanti):
            var = tk.BooleanVar(value=False)
            self._mgmt_check_vars.append(var)

            row_bg = self.C_CARD if i % 2 == 0 else self.C_STRIPE
            row = tk.Frame(self._mgmt_scroll, bg=row_bg)
            row.pack(fill=tk.X, padx=8, pady=6)

            tk.Checkbutton(row, variable=var, bg=row_bg,
                           activebackground=row_bg, cursor="hand2"
                           ).pack(side=tk.LEFT)

            info = tk.Frame(row, bg=row_bg)
            info.pack(side=tk.LEFT, fill=tk.X, expand=True)

            tk.Label(info, text=s["Name"], font=("Segoe UI", 10, "bold"),
                     bg=row_bg, fg="#1E293B", anchor="w").pack(anchor="w")

            host = s["HostAddress"] or "-"
            detail = f"IP/Host: {host}   •   Porta: {s['PortName'] or '-'}"
            if s["Location"]: detail += f"   \U0001f4cd {s['Location']}"
            tk.Label(info, text=detail, font=("Segoe UI", 8),
                     bg=row_bg, fg=self.C_SUBLABEL, anchor="w").pack(anchor="w")

            if s["DriverName"]:
                tk.Label(row, text=s["DriverName"],
                         font=("Segoe UI", 8), bg="#FEF2F2", fg="#DC2626",
                         padx=6, pady=2, relief="flat"
                         ).pack(side=tk.RIGHT, padx=(0, 4))

    def _esporta_installate(self):
        # Esporta la lista attualmente mostrata (già filtrata da "solo rete")
        stampanti = self._mgmt_printers
        if not stampanti:
            messagebox.showwarning(
                "Nessuna stampante",
                "Nessuna stampante da esportare. Premi 'Aggiorna' per ricaricare la lista.",
                parent=self._mgmt_win)
            return

        ts = datetime.datetime.now().strftime("%Y%m%d")
        path = filedialog.asksaveasfilename(
            title="Esporta stampanti installate",
            defaultextension=".csv",
            initialfile=f"stampanti_installate_{ts}.csv",
            filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx")],
            parent=self._mgmt_win
        )
        if not path:
            return

        try:
            esporta_installate(path, stampanti)
        except Exception as e:
            messagebox.showerror("Errore esportazione", str(e), parent=self._mgmt_win)
            return

        self._mgmt_log_write(
            f"\U0001f4be Esportate {len(stampanti)} stampante/i in {os.path.basename(path)}",
            "info")
        messagebox.showinfo(
            "Esportazione completata",
            f"✔ {len(stampanti)} stampante/i esportata/e in:\n{path}\n\n"
            "Nota: le colonne 'cartella_driver' e 'file_inf' (percorso del driver) "
            "sono vuote perché non ricavabili dal sistema.\n"
            "Compilale manualmente prima di reimportare il file su un altro PC.",
            parent=self._mgmt_win)

    def _mgmt_select_all(self):
        for v in self._mgmt_check_vars: v.set(True)

    def _mgmt_select_none(self):
        for v in self._mgmt_check_vars: v.set(False)

    def _avvia_disinstallazione(self):
        if self._mgmt_busy:
            return
        selected = [s for i, s in enumerate(self._mgmt_printers)
                    if i < len(self._mgmt_check_vars) and self._mgmt_check_vars[i].get()]

        if not selected:
            messagebox.showwarning("Nessuna selezione",
                "Seleziona almeno una stampante da disinstallare.",
                parent=self._mgmt_win)
            return

        rm_porta  = self._mgmt_rm_porta.get()
        rm_driver = self._mgmt_rm_driver.get()

        ambito = "la stampante"
        if rm_porta and rm_driver:
            ambito = "la stampante, la porta TCP/IP e il driver (se non più in uso)"
        elif rm_porta:
            ambito = "la stampante e la porta TCP/IP (se non più in uso)"
        elif rm_driver:
            ambito = "la stampante e il driver (se non più in uso)"

        if not messagebox.askyesno(
            "Conferma disinstallazione",
            f"Verranno disinstallate {len(selected)} stampante/i.\n\n"
            f"Per ognuna verrà rimossa: {ambito}.\n\nContinuare?",
            parent=self._mgmt_win
        ):
            return

        self._mgmt_busy = True
        self._mgmt_uninstall_btn.config(
            state="disabled", text="⏳  Disinstallazione in corso...")

        # Logger su file per audit (riusa setup_logger)
        log_dir = os.path.join(
            os.path.dirname(os.path.abspath(sys.argv[0])), "logs")
        self.logger, self.log_file = setup_logger(log_dir)
        self.open_log_btn.config(state="normal")

        threading.Thread(
            target=self._run_disinstallazione,
            args=(selected, rm_porta, rm_driver),
            daemon=True
        ).start()

    def _run_disinstallazione(self, selected: List[dict], rm_porta: bool, rm_driver: bool):
        ts = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        sep = "═" * 44
        self._mgmt_log_write(sep, "head")
        self._mgmt_log_write(f"  {APP_NAME} {APP_VERSION} - DISINSTALLAZIONE", "head")
        self._mgmt_log_write(f"  Avvio: {ts} - {len(selected)} stampante/i", "head")
        self._mgmt_log_write(sep + "\n", "head")

        results = []
        for s in selected:
            try:
                success = disinstalla_stampante(
                    s, rm_porta, rm_driver, self._mgmt_log_write)
            except Exception as e:
                success = False
                self._mgmt_log_write(
                    f"    ✘ Errore inaspettato per {s.get('Name','')}: {e}", "err")
            results.append((s.get("Name", ""), success))

        self.root.after(0, self._fine_disinstallazione, results)

    def _fine_disinstallazione(self, results: list):
        self._mgmt_busy = False
        if self._mgmt_win is not None and self._mgmt_win.winfo_exists():
            self._mgmt_uninstall_btn.config(
                state="normal", text="\U0001f5d1  DISINSTALLA SELEZIONATE")

        ok_list  = [n for n, s in results if s]
        err_list = [n for n, s in results if not s]

        sep = "═" * 44
        self._mgmt_log_write(f"\n{sep}", "head")
        self._mgmt_log_write("  RIEPILOGO DISINSTALLAZIONE", "head")
        self._mgmt_log_write(sep, "head")
        self._mgmt_log_write(f"  ✔ Rimosse : {len(ok_list)}/{len(results)}", "ok")
        if err_list:
            self._mgmt_log_write(f"  ✘ Fallite : {len(err_list)}/{len(results)}", "err")
            for n in err_list:
                self._mgmt_log_write(f"     • {n}", "err")
        if self.log_file:
            self._mgmt_log_write(f"  Log: {self.log_file}", "info")
        self._mgmt_log_write(sep + "\n", "head")

        # Refresh della lista per riflettere le rimozioni
        self._aggiorna_lista_installate()

        parent = self._mgmt_win if (self._mgmt_win and self._mgmt_win.winfo_exists()) else self.root
        if not err_list:
            messagebox.showinfo(
                "Disinstallazione completata",
                f"✔ {len(ok_list)} stampante/i rimossa/e correttamente.",
                parent=parent)
        elif not ok_list:
            messagebox.showerror(
                "Disinstallazione fallita",
                "✘ Nessuna stampante rimossa.\n\nVerifica il log per i dettagli.",
                parent=parent)
        else:
            messagebox.showwarning(
                "Disinstallazione parziale",
                f"✔ Rimosse : {len(ok_list)}\n"
                f"✘ Fallite : {len(err_list)}\n\n"
                + "\n".join(f"  • {n}" for n in err_list),
                parent=parent)


# ─────────────────────────────────────────────────────────────────────────────
#  Schermata di caricamento (splash)
# ─────────────────────────────────────────────────────────────────────────────

def mostra_splash(root: tk.Tk) -> tk.Toplevel:
    """Piccola finestra di caricamento: logo centrato + 'Loading...'.

    Mostrata mentre l'app costruisce la UI e carica l'Excel predefinito, così
    Windows non rileva il programma come 'Non risponde'.
    """
    bg = "#1A2B4A"
    sp = tk.Toplevel(root)
    sp.overrideredirect(True)          # senza bordi/titolo
    sp.configure(bg=bg)
    try:
        sp.attributes("-topmost", True)
    except tk.TclError:
        pass

    # Cornice sottile
    frame = tk.Frame(sp, bg=bg, highlightbackground="#2563EB",
                     highlightthickness=1)
    frame.pack(fill=tk.BOTH, expand=True)

    icon_path = os.path.join(resource_dir(), "assets", "icona.png")
    if _PIL_OK and os.path.isfile(icon_path):
        img = Image.open(icon_path).resize((96, 96), Image.LANCZOS)
        sp._logo_photo = ImageTk.PhotoImage(img, master=sp)
        tk.Label(frame, image=sp._logo_photo, bg=bg).pack(padx=48, pady=(26, 12))
    else:
        tk.Label(frame, text="\U0001f5a8", font=("Segoe UI Emoji", 44),
                 bg=bg, fg="white").pack(padx=48, pady=(26, 12))

    tk.Label(frame, text="Loading...", font=("Segoe UI", 12),
             bg=bg, fg="#CBD5E1").pack(pady=(0, 26))

    # Centra sullo schermo
    sp.update_idletasks()
    w, h = sp.winfo_width(), sp.winfo_height()
    sw, sh = sp.winfo_screenwidth(), sp.winfo_screenheight()
    sp.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
    sp.update()                        # forza il rendering immediato
    return sp


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Controllo UAC prima di aprire la finestra principale
    if not is_admin():
        # Serve un root temporaneo per i messagebox
        _tmp = tk.Tk()
        _tmp.withdraw()
        riavvia = messagebox.askyesno(
            "Privilegi di amministratore richiesti",
            f"{APP_NAME} richiede privilegi di amministratore\n"
            "per installare driver e stampanti di rete.\n\n"
            "Vuoi riavviare il programma con i privilegi necessari?",
            parent=_tmp
        )
        _tmp.destroy()
        if riavvia:
            elevate_and_restart()
        # Se l'utente rifiuta, continua comunque (funzionalità limitate)

    root = tk.Tk()
    root.withdraw()                    # nascosta finché lo splash non finisce
    splash = mostra_splash(root)
    StampaFireApp(root, splash=splash)
    root.mainloop()


if __name__ == "__main__":
    main()
